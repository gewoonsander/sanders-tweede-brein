#!/usr/bin/env python3
"""
foto-doc-verwerker.py

Verwerkt bestanden in een Google Drive map:
- Foto's: analyseren met Ollama vision → hernoemen → importeren in Apple Photos
- Documenten: hernoemen op basis van bestandsnaam via Ollama

Gebruik:
  python3 foto-doc-verwerker.py [mappad] [--recursive]

Standaard map: ~/Library/CloudStorage/GoogleDrive-sander@gewoonsander.nl/Mijn Drive/d archief
"""

import os, sys, json, subprocess, re, base64
from datetime import datetime
from pathlib import Path

def exif_datum(path: Path) -> str:
    """Lees opnamedatum uit EXIF. Geeft 'YYYY-MM-DD-' of '' terug."""
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS
        img = Image.open(path)
        exif = img._getexif()
        if not exif:
            return ""
        for tag_id, waarde in exif.items():
            if TAGS.get(tag_id) in ("DateTimeOriginal", "DateTime"):
                # formaat: "2019:03:15 14:32:00"
                dt = datetime.strptime(waarde[:10], "%Y:%m:%d")
                return dt.strftime("%Y-%m-%d-")
    except Exception:
        pass
    return ""

def bestand_datum(path: Path) -> str:
    """Aanmaak- of wijzigingsdatum van het bestand als 'YYYY-MM-DD-'."""
    try:
        ts = path.stat().st_birthtime  # macOS aanmaakdatum
    except AttributeError:
        ts = path.stat().st_mtime
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d-")

def lees_inhoud(path: Path, max_tekens: int = 600) -> str:
    """Lees tekstinhoud van een document, ongeacht type."""
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            tekst = ""
            for pagina in reader.pages[:3]:
                tekst += pagina.extract_text() or ""
                if len(tekst) >= max_tekens:
                    break
            return tekst[:max_tekens].strip()

        elif ext in {".docx"}:
            from docx import Document
            doc = Document(str(path))
            tekst = " ".join(p.text for p in doc.paragraphs if p.text.strip())
            return tekst[:max_tekens].strip()

        elif ext in {".txt", ".md", ".csv"}:
            return path.read_text(errors="ignore")[:max_tekens].strip()

        elif ext in {".xlsx", ".xls"}:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rijen = []
            for rij in ws.iter_rows(max_row=5, values_only=True):
                rijen.append(" ".join(str(c) for c in rij if c))
            return " | ".join(rijen)[:max_tekens].strip()

    except Exception:
        pass
    return ""

# ── Configuratie ──────────────────────────────────────────────────────────────
OLLAMA_URL   = "http://localhost:11434/api/generate"
VISION_MODEL = "llava:7b"
TEXT_MODEL   = "llama3.2:3b"
DEFAULT_MAP  = Path.home() / "Library/CloudStorage/GoogleDrive-sander@gewoonsander.nl/Mijn Drive/d archief"

FOTO_EXTENSIES = {'.jpg', '.jpeg', '.png', '.gif', '.heic', '.heif',
                  '.raw', '.cr2', '.dng', '.tif', '.tiff', '.webp', '.bmp', '.JPG', '.JPEG', '.PNG'}
SKIP_EXTENSIES = {'.gdoc', '.gsheet', '.gslides', '.gform', '.gdraw',
                  '.zip', '.gz', '.tar'}

# ── Hulpfuncties ─────────────────────────────────────────────────────────────

def is_foto(path: Path) -> bool:
    return path.suffix in FOTO_EXTENSIES

def overslaan(path: Path) -> bool:
    return path.suffix in SKIP_EXTENSIES or path.name.startswith('.')

def opschonen(naam: str) -> str:
    naam = naam.strip().lower()
    naam = re.sub(r'[^a-z0-9\-]', '-', naam)
    naam = re.sub(r'-+', '-', naam).strip('-')
    return naam[:80]

def zeker_datum(naam: str, prefix: str) -> str:
    """Zorg dat de naam altijd begint met het datumprefix."""
    schoon_prefix = prefix.rstrip('-')
    if not naam.startswith(schoon_prefix):
        naam = prefix + naam.lstrip('-')
    return opschonen(naam)

def datumprefix(stem: str) -> str:
    """Haal bestaand datumprefix op als YYYY-MM-DD-, anders lege string."""
    m = re.match(r'^(\d{4}[-_]?\d{2}[-_]?\d{2})', stem)
    if m:
        raw = m.group(1)
        genorm = raw[:4] + '-' + raw[4:6].lstrip('-_') + '-' + raw[6:8].lstrip('-_')
        return genorm.replace('--', '-') + '-'
    return ''

def ollama_verzoek(payload: dict) -> str:
    import urllib.request
    data = json.dumps(payload).encode()
    req = urllib.request.Request(OLLAMA_URL, data=data,
                                  headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=90) as r:
        return json.loads(r.read()).get("response", "").strip()

def beschrijf_foto(path: Path) -> str:
    with open(path, 'rb') as f:
        img_b64 = base64.b64encode(f.read()).decode()

    # Stap 1: llava beschrijft de foto in het Engels
    beschrijving = ollama_verzoek({
        "model": VISION_MODEL,
        "prompt": "Describe the main subject of this photo in one short sentence (max 10 words). Be specific.",
        "images": [img_b64],
        "stream": False
    })

    # Stap 2: llama3.2 vertaalt naar Nederlandse kebab-case naam
    naam = ollama_verzoek({
        "model": TEXT_MODEL,
        "prompt": (
            f"Vertaal naar Nederlandse kebab-case (max 4 woorden, geen lidwoorden de/het/een, alleen kleine letters en koppeltekens).\n"
            f"Voorbeelden:\n"
            f"'elderly man hugging girl' → opa-knuffelt-kleinkind\n"
            f"'baby with ponytails smiling' → baby-paardenstaart-lacht\n"
            f"'bathroom white tiles shower' → badkamer-wit-tegels\n"
            f"Geef ALLEEN het resultaat, geen uitleg.\n"
            f"Vertaal: {beschrijving}"
        ),
        "stream": False
    })
    return naam

def beschrijf_document(path: Path) -> str:
    # Datumprioriteit: bestandsnaam → aanmaakdatum bestand
    prefix = datumprefix(path.stem) or bestand_datum(path)
    inhoud = lees_inhoud(path)

    if inhoud:
        prompt = (
            f"Maak een kebab-case bestandsnaam volgens dit formaat: [type]-[naam]-[onderwerp] (max 6 woorden, kleine letters). "
            f"Gebruik datumprefix: {prefix} "
            f"[type] = documentsoort zoals: brief, factuur, polis, handleiding, offerte, rapport, contract, aangifte. "
            f"[naam] = naam van afzender, ontvanger, organisatie of persoon uit het document (VERPLICHT als aanwezig). "
            f"[onderwerp] = waar het over gaat. "
            f"Voorbeeld: 2021-03-03-brief-tandarts-kerbusch of 2019-10-01-polis-combimotors-motorverzekering. "
            f"Geef ALLEEN de naam terug, zonder uitleg, zonder extensie. "
            f"Bestandsnaam: {path.stem} | Inhoud: {inhoud}"
        )
    else:
        prompt = (
            f"Maak een kebab-case bestandsnaam volgens dit formaat: [type]-[naam]-[onderwerp] (max 6 woorden, kleine letters). "
            f"Gebruik datumprefix: {prefix} "
            f"[type] = documentsoort (brief, factuur, polis, handleiding etc.). "
            f"[naam] = naam van organisatie of persoon als zichtbaar in bestandsnaam. "
            f"Geef ALLEEN de naam terug, zonder uitleg, zonder extensie. "
            f"Bestandsnaam: {path.stem}"
        )
    return ollama_verzoek({"model": TEXT_MODEL, "prompt": prompt, "stream": False})

def importeer_in_apple_photos(path: Path) -> bool:
    script = f'tell application "Photos" to import POSIX file "{path}"'
    result = subprocess.run(["osascript", "-e", script], capture_output=True)
    return result.returncode == 0

# ── Hoofdlogica ───────────────────────────────────────────────────────────────

def verwerk_map(map_pad: Path, recursive: bool = False):
    if recursive:
        bestanden = [f for f in map_pad.rglob('*') if f.is_file()]
    else:
        bestanden = [f for f in map_pad.iterdir() if f.is_file()]

    bestanden = [f for f in bestanden if not overslaan(f)]

    if not bestanden:
        print("Geen bestanden gevonden (of alles overgeslagen).")
        return

    totaal = len(bestanden)
    print(f"\n{totaal} bestanden in '{map_pad.name}/'\n")

    voorstellen = []
    for i, f in enumerate(bestanden, 1):
        print(f"[{i}/{totaal}] {f.name}")
        try:
            if is_foto(f):
                prefix = exif_datum(f) or datumprefix(f.stem) or bestand_datum(f)
                nieuwe_stem = zeker_datum(opschonen(beschrijf_foto(f)), prefix)
                actie = "Apple Photos"
            else:
                prefix = datumprefix(f.stem) or bestand_datum(f)
                nieuwe_stem = zeker_datum(opschonen(beschrijf_document(f)), prefix)
                actie = "Google Drive"

            nieuwe_naam = nieuwe_stem + f.suffix.lower()
            voorstellen.append((f, nieuwe_naam, is_foto(f), actie))
            print(f"        → {nieuwe_naam}  [{actie}]")
        except Exception as e:
            print(f"        FOUT: {e}")
            voorstellen.append((f, None, is_foto(f), "?"))

    print("\n" + "=" * 60)
    print("\nKeuze:")
    print("  A — Alles uitvoeren")
    print("  B — Per stuk bevestigen")
    print("  C — Annuleren")

    keuze = input("\nJouw keuze: ").strip().upper()
    if keuze == 'C':
        print("Geannuleerd.")
        return

    foto_verwijderen = False
    if any(is_foto for _, _, is_foto, _ in voorstellen if is_foto):
        antw = input("\nFoto's na import in Apple Photos verwijderen van Google Drive? (j/n): ").strip().lower()
        foto_verwijderen = antw == 'j'

    for origineel, nieuwe_naam, foto, actie in voorstellen:
        if nieuwe_naam is None:
            print(f"SKIP: {origineel.name}")
            continue

        if keuze == 'B':
            bevestig = input(f"\n{origineel.name} → {nieuwe_naam} [{actie}]? (j/n): ").strip().lower()
            if bevestig != 'j':
                continue

        # Unieke naam bepalen
        nieuw_pad = origineel.parent / nieuwe_naam
        teller = 1
        while nieuw_pad.exists() and nieuw_pad != origineel:
            stam = nieuwe_naam.rsplit('.', 1)[0] + f"-{teller}"
            nieuw_pad = origineel.parent / (stam + origineel.suffix.lower())
            teller += 1

        origineel.rename(nieuw_pad)
        print(f"✓ {origineel.name} → {nieuw_pad.name}")

        if foto:
            if importeer_in_apple_photos(nieuw_pad):
                print(f"  → Geïmporteerd in Apple Photos")
                if foto_verwijderen:
                    nieuw_pad.unlink()
                    print(f"  → Verwijderd van Google Drive")
            else:
                print(f"  → Apple Photos import mislukt (bestand bewaard)")

    print("\nKlaar.")

# ── Entrypoint ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args = sys.argv[1:]
    recursive = "--recursive" in args
    args = [a for a in args if a != "--recursive"]

    map_pad = Path(args[0]).expanduser() if args else DEFAULT_MAP

    if not map_pad.exists():
        print(f"Map niet gevonden: {map_pad}")
        sys.exit(1)

    verwerk_map(map_pad, recursive)
